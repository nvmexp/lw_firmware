#!/usr/bin/python3

# LWIDIA_COPYRIGHT_BEGIN
#
# Copyright 2013, 2017 by LWPU Corporation. All rights reserved. All
# information contained herein is proprietary and confidential to LWPU
# Corporation. Any use, reproduction, or disclosure without the written
# permission of LWPU Corporation is prohibited.
#
# LWIDIA_COPYRIGHT_END 

import argparse, re, sys, os

def QueryYesNo(question, default = "yes"):
    valid = {"yes" : True, "y" : True,  "no" : False, "n":False}
    if default == None:
        prompt = " [y/n] "
    elif default == "yes":
        prompt = " [Y/n] "
    elif default == "no":
        prompt = " [y/N] "
    else:
        raise ValueError("invalid default answer: '%s'" % default)

    while True:
        print(question + prompt)
        choice = input().lower()
        if default is not None and choice == '':
            return valid[default]
        elif choice in valid:
            return valid[choice]
        else:
            print("Please respond with 'yes' or 'no' (or 'y' or 'n').\n")
            
def main():
    pymodulesDir = os.path.join("sw", "mods", "scripts", "pymodules")
    pymodulesDirAbs = os.path.join(os.elwiron["P4ROOT"], pymodulesDir)
    sys.path.append(pymodulesDirAbs)
    
    try:
        from openpyxl import load_workbook
        from openpyxl.cell import column_index_from_string
    except ImportError:
        print("Cannot import openpyxl. Make sure P4ROOT environment variable is" \
              " set properly and {} is in your Perforce client specification." \
              .format(pymodulesDir))
        return 1

    parser = argparse.ArgumentParser()
    parser.add_argument("inXlFile")
    parser.add_argument("inOutJsFile", nargs = '?')
    
    args = parser.parse_args()

    if None != args.inOutJsFile:
        try:
            inOutJsFile = open(args.inOutJsFile, "rt")
            inJs = inOutJsFile.read()
            inOutJsFile.close()
        except IOError as err:
            print(err, file = sys.stderr)
            return 1
    
    wb = load_workbook(filename = args.inXlFile)    
    isSupportedWs = wb.get_sheet_by_name("IsSupported")
    selectedWs = wb.get_sheet_by_name("Selected")
    
    firstRow = min(isSupportedWs.row_dimensions.keys()) - 1
    firstCol = min([column_index_from_string(c) for c in isSupportedWs.column_dimensions.keys()]) - 1
    
    colToBoard = {}
    for c in range(firstCol + 1, isSupportedWs.get_highest_column()):
        colToBoard[c] = str(isSupportedWs.cell(row = firstRow, column = c).value)
        
    colToSpc = {}
    for c in range(firstCol + 1, selectedWs.get_highest_column()):
        colToSpc[c] = str(selectedWs.cell(row = firstRow, column = c).value)
    
    testsInExcel = {}
    for r in range(firstRow + 1, isSupportedWs.get_highest_row()):
        sup = []
        testName = str(isSupportedWs.cell(row = r, column = 0).value)
        for c in range(firstCol + 1, isSupportedWs.get_highest_column()):
            if 1 == isSupportedWs.cell(row = r, column = c).value:
                sup.append(colToBoard[c])
        sel = {}
        for c in range(firstCol + 1, selectedWs.get_highest_column()):
            sel[colToSpc[c]] = int(selectedWs.cell(row = r, column = c).value)
        testsInExcel[testName] = {"Sel" : sel, "Sup" : sup}
        
    testNameWidth = max([len(testName) for testName in testsInExcel.keys()]) + 1
    testsSelectedAndSupportedTableInExcel = "g_TestsSelectedAndSupported =\n{\n"
    firstTestIt = True
    for (k, v) in sorted(testsInExcel.items()):
        if firstTestIt:
            firstTestIt = False
        else:
            testsSelectedAndSupportedTableInExcel += ",//$\n"
        testsSelectedAndSupportedTableInExcel += "    {:<{testNameWidth}}: {{Sel: {{dvs" \
            ":{dvs}, mfg:{mfg}, compute:{compute}, oqa:{oqa}, cheetah:{cheetah}}}, "\
            .format(
                k,
                testNameWidth = testNameWidth,
                dvs      = v["Sel"]["dvs"],
                mfg      = v["Sel"]["mfg"],
                compute  = v["Sel"]["compute"],
                oqa      = v["Sel"]["oqa"],
                cheetah    = v["Sel"]["cheetah"])
        supArray = "["
        for (i, s) in enumerate(v["Sup"]):
            if 0 != i: supArray += ", "
            supArray += '"{}"'.format(s)
        supArray += "]"
        testsSelectedAndSupportedTableInExcel += "Sup: {} }}".format(supArray)
            
    testsSelectedAndSupportedTableInExcel += "//$\n}"

    if None != args.inOutJsFile:
        testsSelectedAndSupportedTableRe = r"g_TestsSelectedAndSupported\s*=\s*{" \
            "\s*(?P<tests>(\w+\s*:\s*{Sel\s*:\s*{dvs\s*:\s*[01]\s*,\s*mfg\s*:\s*" \
            "[01]\s*,\s*compute\s*:\s*[01]\s*,\s*oqa\s*:\s*[01]\s*,\s*cheetah\s*:" \
            "\s*[01]\s*}\s*,\s*Sup\s*:\s*\[\s*([^]]*)\s*\]\s*}\s*,?\s*(//[^\n]*$)?" \
            "\s*)*)\s*}" 
        m = re.search(testsSelectedAndSupportedTableRe, inJs, re.DOTALL | re.MULTILINE)
        if None == m:
            print("Cannot find g_TestsSelectedAndSupported or its format is not supported", file = sys.stderr)
            return 1
        testsSelectedAndSupportedTable = m.group("tests")
        oneTestRe = r"(?P<testName>\w+)\s*:\s*{Sel\s*:\s*{dvs\s*\s*:\s*(?P<dvs>" \
            "[01])\s*,\s*mfg\s*:\s*(?P<mfg>[01])\s*,\s*compute\s*:\s*(?P<compute>" \
            "[01])\s*,\s*oqa\s*:\s*(?P<oqa>[01])\s*,\s*cheetah\s*:\s*(?P<cheetah>[01])" \
            "\s*}\s*,\s*Sup\s*:\s*(?P<boards>\[\s*[^]]*\s*\])\s*}"
        testsInJs = {}
        spcs = ["dvs", "mfg", "compute", "oqa", "cheetah"]
        
        for t in re.finditer(oneTestRe, testsSelectedAndSupportedTable):
            testSel = {}
            for s in spcs:
                testSel[s] = int(t.group(s))
            testSup = eval(t.group("boards"))
            testSup.sort()
            testsInJs[t.group("testName")] = {"Sel" : testSel, "Sup" : testSup}
        testsInJsSet = set(testsInJs.keys())
        testsInExcelSet = set(testsInExcel.keys())
        added = testsInExcelSet - testsInJsSet
        removed = testsInJsSet - testsInExcelSet
        inBoth = testsInExcelSet.intersection(testsInJsSet)
        changed = False
        if set() != added:
            changed = True
            print("The following tests were added:")
            for t in added:
                print("    {}: {}".format(t, testsInExcel[t]))  
        if set() != removed:
            changed = True
            print("The following tests were removed:")
            for t in removed:
                print("    {}: {}".format(t, testsInJs[t]))
        modified = []
        for t in inBoth:
            if testsInExcel[t] != testsInJs[t]:
                modified.append(t)
        if len(modified) != 0:
            changed = True
            print("The following tests were modified:")
            for t in modified:
                print("    {}".format(t))
                print("        old: {}".format(testsInJs[t]))
                print("        new: {}".format(testsInExcel[t]))
        if not changed:
            print('The JavaScript file and the spreadsheet are equivalent. No changes to "{}" will be written.'.format(args.inOutJsFile))
        else:
            if QueryYesNo("Proceed with changes?"):
                outJs = re.sub(testsSelectedAndSupportedTableRe, testsSelectedAndSupportedTableInExcel, inJs, flags = re.DOTALL | re.MULTILINE)
                inOutJsFile = open(args.inOutJsFile, "wt")
                inOutJsFile.write(outJs)
    else:
        print("{}".format(testsSelectedAndSupportedTableInExcel))
    
    return 0

if __name__ == '__main__':
    exit(main())
