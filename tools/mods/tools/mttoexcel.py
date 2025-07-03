#!/usr/bin/python3

# LWIDIA_COPYRIGHT_BEGIN
#
# Copyright 2013 by LWPU Corporation. All rights reserved. All information
# contained herein is proprietary and confidential to LWPU Corporation. Any
# use, reproduction, or disclosure without the written permission of LWPU
# Corporation is prohibited.
#
# LWIDIA_COPYRIGHT_END 

import argparse, re, sys, os

def main():
    pymodulesDir = os.path.join("sw", "mods", "scripts", "pymodules")
    pymodulesDirAbs = os.path.join(os.elwiron["P4ROOT"], pymodulesDir)
    sys.path.append(pymodulesDirAbs)
    
    try:
        from openpyxl import Workbook
        from openpyxl.datavalidation import DataValidation, ValidationType
    except ImportError:
        print("Cannot import openpyxl. Make sure P4ROOT environment variable is" \
              " set properly and {} is in your Perforce client specification." \
              .format(pymodulesDir))
        return 1

    parser = argparse.ArgumentParser()
    parser.add_argument("inJsFile")
    parser.add_argument("outExcelFile")
    
    args = parser.parse_args()
    
    inJs = open(args.inJsFile, "rt").read()
    
    testsSelectedAndSupportedTableRe = r"g_TestsSelectedAndSupported\s*=\s*{" \
        "\s*(?P<tests>(\w+\s*:\s*{Sel\s*:\s*{dvs\s*:\s*[01]\s*,\s*mfg\s*:\s*" \
        "[01]\s*,\s*compute\s*:\s*[01]\s*,\s*oqa\s*:\s*[01]\s*,\s*cheetah\s*:" \
        "\s*[01]\s*}\s*,\s*Sup\s*:\s*\[\s*([^]]*)\s*\]\s*}\s*,?\s*(//[^\n]*$)?" \
        "\s*)*)\s*}" 

    m = re.search(testsSelectedAndSupportedTableRe, inJs, re.DOTALL | re.MULTILINE)
    if None == m:
        print("Cannot find g_TestsSelectedAndSupported or its format is not supported", file = sys.stderr)
        return 1
    
    testsSelectedAndSupportedTable = m.group("tests")
    oneTestRe = r"(?P<testName>\w+)\s*:\s*{Sel\s*:\s*{dvs\s*\s*:\s*(?P<dvs>" \
        "[01])\s*,\s*mfg\s*:\s*(?P<mfg>[01])\s*,\s*compute\s*:\s*(?P<compute>" \
        "[01])\s*,\s*oqa\s*:\s*(?P<oqa>[01])\s*,\s*cheetah\s*:\s*(?P<cheetah>[01])" \
        "\s*}\s*,\s*Sup\s*:\s*(?P<boards>\[\s*[^]]*\s*\])\s*}"
        
    tests = {}

    spcs = ["dvs", "mfg", "compute", "oqa", "cheetah"]
    
    for t in re.finditer(oneTestRe, testsSelectedAndSupportedTable):
        testSel = {}
        for s in spcs:
            testSel[s] = int(t.group(s))
        testSup = eval(t.group("boards"))
        tests[t.group("testName")] = {"Sel" : testSel, "Sup" : testSup}
        
    boards = set([])
    for v in tests.values():
        boards.update(v["Sup"])
    boardToCol = {}
    for (i, b) in enumerate(sorted(boards)):
        boardToCol[b] = i + 1
    spcToCol = {}
    for (i, s) in enumerate(spcs):
        spcToCol[s] = i + 1;  

    dv1 = DataValidation(ValidationType.LIST, formula1 = '"0,1"', allow_blank = False)
    
    wb = Workbook()
    ws = wb.get_active_sheet()
    ws.title = "IsSupported"
    ws.freeze_panes = "A2"
    ws.add_data_validation(dv1)
    for (k, v) in boardToCol.items():
        ws.cell(row = 0, column = v).value = k
    for (i, (k, v)) in enumerate(sorted(tests.items())):
        ws.cell(row = 1 + i, column = 0).value = k
        for col in range(1, 1 + len(boards)):
            c = ws.cell(row = 1 + i, column = col)
            c.value = 0
            dv1.add_cell(c)
        for b in v["Sup"]:
            ws.cell(row = 1 + i, column = boardToCol[b]).value = 1
    ws.column_dimensions["A"].width = 23
    
    ws = wb.create_sheet(title = "Selected")
    ws.freeze_panes = "A2"
    dv2 = DataValidation(ValidationType.LIST, formula1 = '"0,1"', allow_blank = False)
    ws.add_data_validation(dv2)
    for (k, v) in spcToCol.items():
        ws.cell(row = 0, column = v).value = k
    for (i, (k, v)) in enumerate(sorted(tests.items())):
        ws.cell(row = 1 + i, column = 0).value = "=IsSupported!A" + str(i + 2)
        for col in range(1, 1 + len(spcs)):
            c = ws.cell(row = 1 + i, column = col)
            c.value = 0
            dv2.add_cell(c)
        for (spc, isSel) in v["Sel"].items():
            ws.cell(row = 1 + i, column = spcToCol[spc]).value = isSel
    ws.column_dimensions["A"].width = 23
    
    wb.save(args.outExcelFile)
    
    return 0

if __name__ == '__main__':
    exit(main())